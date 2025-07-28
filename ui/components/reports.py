#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Composant rapports et analyses
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, date, timedelta
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import sqlite3

class ReportsFrame(ttk.Frame):
    """Frame pour les rapports et analyses"""
    
    def __init__(self, parent, db_manager, current_user):
        super().__init__(parent)
        self.db = db_manager
        self.current_user = current_user
        
        # Interface utilisateur
        self.setup_ui()
        
        # Charger les données initiales
        self.load_initial_data()
    
    def setup_ui(self):
        """Configure l'interface utilisateur"""
        # Titre
        title_frame = ttk.Frame(self)
        title_frame.pack(fill=tk.X, padx=20, pady=10)
        
        ttk.Label(title_frame, text="📈 Rapports et Analyses", 
                 font=('Arial', 16, 'bold')).pack(side=tk.LEFT)
        
        ttk.Button(title_frame, text="🔄 Actualiser", 
                  command=self.refresh_reports).pack(side=tk.RIGHT)
        
        # Notebook pour les différents rapports
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        # Onglets
        self.create_summary_tab()
        self.create_temporal_tab()
        self.create_banks_tab()
        self.create_clients_tab()
        self.create_export_tab()
    
    def create_summary_tab(self):
        """Crée l'onglet résumé général"""
        summary_frame = ttk.Frame(self.notebook)
        self.notebook.add(summary_frame, text="📊 Résumé Général")
        
        # Période de sélection
        period_frame = ttk.LabelFrame(summary_frame, text="Période d'Analyse", padding=10)
        period_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Label(period_frame, text="Du:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
        from tkcalendar import DateEntry
        self.summary_date_from = DateEntry(period_frame, width=12, background='darkblue',
                                          foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy')
        self.summary_date_from.grid(row=0, column=1, sticky=tk.W, padx=(0, 20))
        
        ttk.Label(period_frame, text="Au:").grid(row=0, column=2, sticky=tk.W, padx=(0, 10))
        self.summary_date_to = DateEntry(period_frame, width=12, background='darkblue',
                                        foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy')
        self.summary_date_to.grid(row=0, column=3, sticky=tk.W, padx=(0, 20))
        
        ttk.Button(period_frame, text="📊 Générer", 
                  command=self.generate_summary).grid(row=0, column=4, padx=10)
        
        # Initialiser les dates
        today = date.today()
        self.summary_date_from.set_date(today - timedelta(days=30))
        self.summary_date_to.set_date(today)
        
        # Zone des statistiques
        stats_frame = ttk.LabelFrame(summary_frame, text="Statistiques", padding=10)
        stats_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Créer les cartes de statistiques
        self.create_summary_cards(stats_frame)
        
        # Zone des graphiques
        charts_frame = ttk.Frame(summary_frame)
        charts_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Graphique des statuts
        status_frame = ttk.LabelFrame(charts_frame, text="Répartition par Statut", padding=10)
        status_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        
        self.summary_status_chart = ttk.Frame(status_frame)
        self.summary_status_chart.pack(fill=tk.BOTH, expand=True)
        
        # Graphique des montants
        amount_frame = ttk.LabelFrame(charts_frame, text="Évolution des Montants", padding=10)
        amount_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(5, 0))
        
        self.summary_amount_chart = ttk.Frame(amount_frame)
        self.summary_amount_chart.pack(fill=tk.BOTH, expand=True)
    
    def create_summary_cards(self, parent):
        """Crée les cartes de statistiques du résumé"""
        cards_frame = ttk.Frame(parent)
        cards_frame.pack(fill=tk.X, pady=10)
        
        # Configuration des cartes
        self.summary_cards = {}
        card_configs = [
            ("total_cheques", "📊 Total Chèques", "0"),
            ("total_amount", "💰 Montant Total", "0 MAD"),
            ("avg_amount", "📈 Montant Moyen", "0 MAD"),
            ("pending_count", "⏳ En Attente", "0")
        ]
        
        for i, (key, title, default_value) in enumerate(card_configs):
            card_frame = ttk.LabelFrame(cards_frame, text=title, padding=10)
            card_frame.grid(row=0, column=i, padx=5, pady=5, sticky="ew")
            
            value_label = ttk.Label(card_frame, text=default_value, 
                                   font=('Arial', 14, 'bold'))
            value_label.pack()
            
            self.summary_cards[key] = value_label
            cards_frame.columnconfigure(i, weight=1)
    
    def create_temporal_tab(self):
        """Crée l'onglet analyse temporelle"""
        temporal_frame = ttk.Frame(self.notebook)
        self.notebook.add(temporal_frame, text="📅 Analyse Temporelle")
        
        # Contrôles
        controls_frame = ttk.Frame(temporal_frame)
        controls_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Label(controls_frame, text="Type d'analyse:").pack(side=tk.LEFT)
        
        self.temporal_type_var = tk.StringVar(value="monthly")
        ttk.Radiobutton(controls_frame, text="Mensuelle", variable=self.temporal_type_var, 
                       value="monthly", command=self.update_temporal_chart).pack(side=tk.LEFT, padx=(10, 0))
        ttk.Radiobutton(controls_frame, text="Hebdomadaire", variable=self.temporal_type_var, 
                       value="weekly", command=self.update_temporal_chart).pack(side=tk.LEFT, padx=(10, 0))
        ttk.Radiobutton(controls_frame, text="Quotidienne", variable=self.temporal_type_var, 
                       value="daily", command=self.update_temporal_chart).pack(side=tk.LEFT, padx=(10, 0))
        
        # Zone du graphique
        chart_frame = ttk.LabelFrame(temporal_frame, text="Évolution Temporelle", padding=10)
        chart_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        self.temporal_chart_frame = ttk.Frame(chart_frame)
        self.temporal_chart_frame.pack(fill=tk.BOTH, expand=True)
    
    def create_banks_tab(self):
        """Crée l'onglet analyse par banque"""
        banks_frame = ttk.Frame(self.notebook)
        self.notebook.add(banks_frame, text="🏦 Analyse par Banque")
        
        # Tableau des banques
        table_frame = ttk.LabelFrame(banks_frame, text="Performance par Banque", padding=10)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Treeview
        columns = ('Banque', 'Nb Chèques', 'Montant Total', 'Montant Moyen', 'Encaissés', 'Rejetés', 'Part %')
        self.banks_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=12)
        
        for col in columns:
            self.banks_tree.heading(col, text=col)
            self.banks_tree.column(col, width=120, anchor=tk.CENTER)
        
        # Scrollbar
        banks_scroll = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.banks_tree.yview)
        self.banks_tree.configure(yscrollcommand=banks_scroll.set)
        
        self.banks_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        banks_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Graphique des banques
        bank_chart_frame = ttk.LabelFrame(banks_frame, text="Répartition par Banque", padding=10)
        bank_chart_frame.pack(fill=tk.X, padx=10, pady=10, ipady=200)
        
        self.bank_chart_frame = ttk.Frame(bank_chart_frame)
        self.bank_chart_frame.pack(fill=tk.BOTH, expand=True)
    
    def create_clients_tab(self):
        """Crée l'onglet analyse des clients"""
        clients_frame = ttk.Frame(self.notebook)
        self.notebook.add(clients_frame, text="👥 Analyse Clients")
        
        # Top clients
        top_frame = ttk.LabelFrame(clients_frame, text="Top 20 Clients", padding=10)
        top_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        columns = ('Rang', 'Client', 'Type', 'Nb Chèques', 'Montant Total', 'Montant Moyen', 'Dernière Activité')
        self.clients_tree = ttk.Treeview(top_frame, columns=columns, show='headings', height=15)
        
        for col in columns:
            self.clients_tree.heading(col, text=col)
            if col == 'Client':
                self.clients_tree.column(col, width=200)
            else:
                self.clients_tree.column(col, width=100, anchor=tk.CENTER)
        
        # Scrollbar
        clients_scroll = ttk.Scrollbar(top_frame, orient=tk.VERTICAL, command=self.clients_tree.yview)
        self.clients_tree.configure(yscrollcommand=clients_scroll.set)
        
        self.clients_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        clients_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Statistiques clients
        client_stats_frame = ttk.LabelFrame(clients_frame, text="Statistiques Clients", padding=10)
        client_stats_frame.pack(fill=tk.X, padx=10, pady=10)
        
        self.client_stats_text = tk.Text(client_stats_frame, height=6, wrap=tk.WORD)
        self.client_stats_text.pack(fill=tk.X)
    
    def create_export_tab(self):
        """Crée l'onglet export des rapports"""
        export_frame = ttk.Frame(self.notebook)
        self.notebook.add(export_frame, text="📤 Export")
        
        # Options d'export
        options_frame = ttk.LabelFrame(export_frame, text="Options d'Export", padding=15)
        options_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # Type de rapport
        ttk.Label(options_frame, text="Type de rapport:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.export_type_var = tk.StringVar(value="summary")
        
        export_types = [
            ("summary", "Résumé général"),
            ("temporal", "Analyse temporelle"),
            ("banks", "Analyse par banque"),
            ("clients", "Analyse clients"),
            ("custom", "Rapport personnalisé")
        ]
        
        for i, (value, text) in enumerate(export_types):
            ttk.Radiobutton(options_frame, text=text, variable=self.export_type_var, 
                           value=value).grid(row=i+1, column=0, sticky=tk.W, padx=20)
        
        # Format d'export
        ttk.Label(options_frame, text="Format:").grid(row=0, column=1, sticky=tk.W, padx=(50, 0), pady=5)
        self.export_format_var = tk.StringVar(value="pdf")
        
        formats = [("pdf", "PDF"), ("xlsx", "Excel"), ("csv", "CSV")]
        for i, (value, text) in enumerate(formats):
            ttk.Radiobutton(options_frame, text=text, variable=self.export_format_var, 
                           value=value).grid(row=i+1, column=1, sticky=tk.W, padx=(70, 0))
        
        # Période pour l'export
        period_frame = ttk.LabelFrame(export_frame, text="Période", padding=15)
        period_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Label(period_frame, text="Du:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
        self.export_date_from = DateEntry(period_frame, width=12, background='darkblue',
                                         foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy')
        self.export_date_from.grid(row=0, column=1, sticky=tk.W, padx=(0, 20))
        
        ttk.Label(period_frame, text="Au:").grid(row=0, column=2, sticky=tk.W, padx=(0, 10))
        self.export_date_to = DateEntry(period_frame, width=12, background='darkblue',
                                       foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy')
        self.export_date_to.grid(row=0, column=3, sticky=tk.W)
        
        # Boutons d'export
        buttons_frame = ttk.Frame(export_frame)
        buttons_frame.pack(fill=tk.X, padx=10, pady=20)
        
        ttk.Button(buttons_frame, text="📊 Aperçu", 
                  command=self.preview_report).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(buttons_frame, text="📤 Exporter", 
                  command=self.export_report).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(buttons_frame, text="📧 Envoyer par Email", 
                  command=self.email_report).pack(side=tk.LEFT)
    
    def load_initial_data(self):
        """Charge les données initiales"""
        self.generate_summary()
        self.update_temporal_chart()
        self.update_banks_analysis()
        self.update_clients_analysis()
    
    def generate_summary(self):
        """Génère le résumé général"""
        try:
            # Récupérer les dates
            date_from = self.summary_date_from.get_date().strftime('%Y-%m-%d')
            date_to = self.summary_date_to.get_date().strftime('%Y-%m-%d')
            
            # Requête pour les statistiques
            with sqlite3.connect(self.db.db_path) as conn:
                cursor = conn.cursor()
                
                # Statistiques générales
                cursor.execute("""
                    SELECT 
                        COUNT(*) as total_cheques,
                        COALESCE(SUM(amount), 0) as total_amount,
                        COALESCE(AVG(amount), 0) as avg_amount,
                        SUM(CASE WHEN status = 'en_attente' THEN 1 ELSE 0 END) as pending_count
                    FROM cheques 
                    WHERE due_date BETWEEN ? AND ?
                """, (date_from, date_to))
                
                stats = cursor.fetchone()
                
                # Mettre à jour les cartes
                self.summary_cards['total_cheques'].config(text=str(stats[0]))
                self.summary_cards['total_amount'].config(text=f"{stats[1]:,.0f} MAD")
                self.summary_cards['avg_amount'].config(text=f"{stats[2]:,.0f} MAD")
                self.summary_cards['pending_count'].config(text=str(stats[3]))
                
                # Données pour le graphique des statuts
                cursor.execute("""
                    SELECT status, COUNT(*), SUM(amount)
                    FROM cheques 
                    WHERE due_date BETWEEN ? AND ?
                    GROUP BY status
                """, (date_from, date_to))
                
                status_data = cursor.fetchall()
                self.update_summary_status_chart(status_data)
                
                # Données pour le graphique des montants
                cursor.execute("""
                    SELECT 
                        strftime('%Y-%m', due_date) as month,
                        SUM(amount) as total
                    FROM cheques 
                    WHERE due_date BETWEEN ? AND ?
                    GROUP BY strftime('%Y-%m', due_date)
                    ORDER BY month
                """, (date_from, date_to))
                
                amount_data = cursor.fetchall()
                self.update_summary_amount_chart(amount_data)
                
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la génération du résumé: {e}")
    
    def update_summary_status_chart(self, data):
        """Met à jour le graphique des statuts du résumé"""
        # Nettoyer le frame
        for widget in self.summary_status_chart.winfo_children():
            widget.destroy()
        
        if not data:
            ttk.Label(self.summary_status_chart, text="Aucune donnée").pack()
            return
        
        # Créer le graphique
        fig, ax = plt.subplots(figsize=(5, 4))
        
        labels = []
        sizes = []
        colors = []
        
        status_colors = {
            'en_attente': '#ffc107',
            'encaisse': '#28a745',
            'rejete': '#dc3545',
            'impaye': '#fd7e14',
            'depose': '#17a2b8',
            'annule': '#6c757d'
        }
        
        status_labels = {
            'en_attente': 'En Attente',
            'encaisse': 'Encaissé',
            'rejete': 'Rejeté',
            'impaye': 'Impayé',
            'depose': 'Déposé',
            'annule': 'Annulé'
        }
        
        for status, count, amount in data:
            labels.append(status_labels.get(status, status))
            sizes.append(count)
            colors.append(status_colors.get(status, '#6c757d'))
        
        ax.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', startangle=90)
        ax.set_title('Répartition par Statut')
        
        # Intégrer dans Tkinter
        canvas = FigureCanvasTkAgg(fig, self.summary_status_chart)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
        plt.close(fig)
    
    def update_summary_amount_chart(self, data):
        """Met à jour le graphique des montants du résumé"""
        # Nettoyer le frame
        for widget in self.summary_amount_chart.winfo_children():
            widget.destroy()
        
        if not data:
            ttk.Label(self.summary_amount_chart, text="Aucune donnée").pack()
            return
        
        # Créer le graphique
        fig, ax = plt.subplots(figsize=(5, 4))
        
        months = [row[0] for row in data]
        amounts = [row[1] for row in data]
        
        ax.bar(months, amounts, color='#007bff', alpha=0.7)
        ax.set_title('Évolution des Montants')
        ax.set_ylabel('Montant (MAD)')
        ax.tick_params(axis='x', rotation=45)
        
        plt.tight_layout()
        
        # Intégrer dans Tkinter
        canvas = FigureCanvasTkAgg(fig, self.summary_amount_chart)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
        plt.close(fig)
    
    def update_temporal_chart(self):
        """Met à jour le graphique temporel"""
        # Nettoyer le frame
        for widget in self.temporal_chart_frame.winfo_children():
            widget.destroy()
        
        try:
            analysis_type = self.temporal_type_var.get()
            
            with sqlite3.connect(self.db.db_path) as conn:
                cursor = conn.cursor()
                
                if analysis_type == "monthly":
                    cursor.execute("""
                        SELECT 
                            strftime('%Y-%m', due_date) as period,
                            COUNT(*) as count,
                            SUM(amount) as total
                        FROM cheques 
                        WHERE due_date >= date('now', '-12 months')
                        GROUP BY strftime('%Y-%m', due_date)
                        ORDER BY period
                    """)
                elif analysis_type == "weekly":
                    cursor.execute("""
                        SELECT 
                            strftime('%Y-W%W', due_date) as period,
                            COUNT(*) as count,
                            SUM(amount) as total
                        FROM cheques 
                        WHERE due_date >= date('now', '-12 weeks')
                        GROUP BY strftime('%Y-W%W', due_date)
                        ORDER BY period
                    """)
                else:  # daily
                    cursor.execute("""
                        SELECT 
                            date(due_date) as period,
                            COUNT(*) as count,
                            SUM(amount) as total
                        FROM cheques 
                        WHERE due_date >= date('now', '-30 days')
                        GROUP BY date(due_date)
                        ORDER BY period
                    """)
                
                data = cursor.fetchall()
                
                if not data:
                    ttk.Label(self.temporal_chart_frame, text="Aucune donnée disponible").pack()
                    return
                
                # Créer le graphique
                fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(10, 6), sharex=True)
                
                periods = [row[0] for row in data]
                counts = [row[1] for row in data]
                amounts = [row[2] for row in data]
                
                # Graphique du nombre
                ax1.plot(periods, counts, marker='o', color='#007bff', linewidth=2)
                ax1.set_ylabel('Nombre de chèques')
                ax1.set_title(f'Évolution {analysis_type.title()}')
                ax1.grid(True, alpha=0.3)
                
                # Graphique des montants
                ax2.plot(periods, amounts, marker='s', color='#28a745', linewidth=2)
                ax2.set_ylabel('Montant (MAD)')
                ax2.set_xlabel('Période')
                ax2.grid(True, alpha=0.3)
                
                # Rotation des labels
                plt.xticks(rotation=45)
                plt.tight_layout()
                
                # Intégrer dans Tkinter
                canvas = FigureCanvasTkAgg(fig, self.temporal_chart_frame)
                canvas.draw()
                canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
                
                plt.close(fig)
                
        except Exception as e:
            ttk.Label(self.temporal_chart_frame, 
                     text=f"Erreur: {e}").pack()
    
    def update_banks_analysis(self):
        """Met à jour l'analyse par banque"""
        try:
            with sqlite3.connect(self.db.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT 
                        bk.name as bank_name,
                        COUNT(c.id) as cheque_count,
                        COALESCE(SUM(c.amount), 0) as total_amount,
                        COALESCE(AVG(c.amount), 0) as avg_amount,
                        SUM(CASE WHEN c.status = 'encaisse' THEN 1 ELSE 0 END) as encaisse_count,
                        SUM(CASE WHEN c.status = 'rejete' THEN 1 ELSE 0 END) as rejete_count
                    FROM banks bk
                    LEFT JOIN branches b ON bk.id = b.bank_id
                    LEFT JOIN cheques c ON b.id = c.branch_id
                    WHERE bk.active = TRUE
                    GROUP BY bk.id, bk.name
                    ORDER BY total_amount DESC
                """)
                
                banks_data = cursor.fetchall()
                
                # Calculer le total pour les pourcentages
                total_amount = sum(row[2] for row in banks_data)
                
                # Vider le tree
                for item in self.banks_tree.get_children():
                    self.banks_tree.delete(item)
                
                # Ajouter les données
                chart_data = []
                for bank_data in banks_data:
                    bank_name, count, amount, avg, encaisse, rejete = bank_data
                    percentage = (amount / total_amount * 100) if total_amount > 0 else 0
                    
                    self.banks_tree.insert('', tk.END, values=(
                        bank_name,
                        count,
                        f"{amount:,.0f}",
                        f"{avg:,.0f}",
                        encaisse,
                        rejete,
                        f"{percentage:.1f}%"
                    ))
                    
                    if amount > 0:
                        chart_data.append((bank_name, amount))
                
                # Mettre à jour le graphique
                self.update_bank_chart(chart_data)
                
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'analyse des banques: {e}")
    
    def update_bank_chart(self, data):
        """Met à jour le graphique des banques"""
        # Nettoyer le frame
        for widget in self.bank_chart_frame.winfo_children():
            widget.destroy()
        
        if not data:
            ttk.Label(self.bank_chart_frame, text="Aucune donnée").pack()
            return
        
        # Créer le graphique
        fig, ax = plt.subplots(figsize=(10, 4))
        
        banks = [row[0] for row in data[:10]]  # Top 10
        amounts = [row[1] for row in data[:10]]
        
        bars = ax.bar(banks, amounts, color='#007bff', alpha=0.7)
        ax.set_title('Top 10 Banques par Montant')
        ax.set_ylabel('Montant (MAD)')
        
        # Ajouter les valeurs sur les barres
        for bar, amount in zip(bars, amounts):
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height,
                   f'{amount:,.0f}', ha='center', va='bottom', fontsize=8)
        
        plt.xticks(rotation=45)
        plt.tight_layout()
        
        # Intégrer dans Tkinter
        canvas = FigureCanvasTkAgg(fig, self.bank_chart_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
        plt.close(fig)
    
    def update_clients_analysis(self):
        """Met à jour l'analyse des clients"""
        try:
            with sqlite3.connect(self.db.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT 
                        cl.name,
                        cl.type,
                        COUNT(c.id) as cheque_count,
                        COALESCE(SUM(c.amount), 0) as total_amount,
                        COALESCE(AVG(c.amount), 0) as avg_amount,
                        MAX(c.created_at) as last_activity
                    FROM clients cl
                    LEFT JOIN cheques c ON cl.id = c.client_id
                    WHERE cl.active = TRUE
                    GROUP BY cl.id, cl.name, cl.type
                    HAVING cheque_count > 0
                    ORDER BY total_amount DESC
                    LIMIT 20
                """)
                
                clients_data = cursor.fetchall()
                
                # Vider le tree
                for item in self.clients_tree.get_children():
                    self.clients_tree.delete(item)
                
                # Ajouter les données
                for i, client_data in enumerate(clients_data, 1):
                    name, client_type, count, amount, avg, last_activity = client_data
                    
                    # Formater la date
                    if last_activity:
                        try:
                            last_date = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S').strftime('%d/%m/%Y')
                        except:
                            last_date = last_activity[:10]
                    else:
                        last_date = "N/A"
                    
                    type_display = "👤 Personne" if client_type == "personne" else "🏢 Entreprise"
                    
                    self.clients_tree.insert('', tk.END, values=(
                        f"#{i}",
                        name,
                        type_display,
                        count,
                        f"{amount:,.0f}",
                        f"{avg:,.0f}",
                        last_date
                    ))
                
                # Statistiques générales des clients
                cursor.execute("""
                    SELECT 
                        COUNT(DISTINCT cl.id) as total_clients,
                        SUM(CASE WHEN cl.type = 'personne' THEN 1 ELSE 0 END) as personnes,
                        SUM(CASE WHEN cl.type = 'entreprise' THEN 1 ELSE 0 END) as entreprises,
                        COUNT(c.id) as total_cheques,
                        COALESCE(AVG(client_cheques.cheque_count), 0) as avg_cheques_per_client
                    FROM clients cl
                    LEFT JOIN cheques c ON cl.id = c.client_id
                    LEFT JOIN (
                        SELECT client_id, COUNT(*) as cheque_count
                        FROM cheques
                        GROUP BY client_id
                    ) client_cheques ON cl.id = client_cheques.client_id
                    WHERE cl.active = TRUE
                """)
                
                stats = cursor.fetchone()
                
                # Afficher les statistiques
                stats_text = f"""
📊 STATISTIQUES CLIENTS

👥 Total clients: {stats[0]}
👤 Personnes physiques: {stats[1]}
🏢 Entreprises: {stats[2]}
📋 Total chèques: {stats[3]}
📈 Moyenne chèques/client: {stats[4]:.1f}

💡 Les 20 meilleurs clients représentent {len(clients_data)} clients actifs.
                """.strip()
                
                self.client_stats_text.delete(1.0, tk.END)
                self.client_stats_text.insert(1.0, stats_text)
                
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'analyse des clients: {e}")
    
    def refresh_reports(self):
        """Actualise tous les rapports"""
        self.load_initial_data()
    
    def show_export_dialog(self):
        """Affiche la boîte de dialogue d'export"""
        self.notebook.select(4)  # Onglet export
    
    def preview_report(self):
        """Aperçu du rapport sélectionné"""
        report_type = self.export_type_var.get()
        messagebox.showinfo("Aperçu", f"Aperçu du rapport '{report_type}' à implémenter")
    
    def export_report(self):
        """Exporte le rapport sélectionné"""
        report_type = self.export_type_var.get()
        format_type = self.export_format_var.get()
        
        # Sélectionner le fichier de destination
        file_types = {
            'pdf': [("PDF files", "*.pdf")],
            'xlsx': [("Excel files", "*.xlsx")],
            'csv': [("CSV files", "*.csv")]
        }
        
        filename = filedialog.asksaveasfilename(
            title=f"Exporter le rapport {report_type}",
            defaultextension=f".{format_type}",
            filetypes=file_types.get(format_type, [("All files", "*.*")])
        )
        
        if filename:
            try:
                if format_type == 'pdf':
                    self.export_to_pdf(report_type, filename)
                elif format_type == 'xlsx':
                    self.export_to_excel(report_type, filename)
                elif format_type == 'csv':
                    self.export_to_csv(report_type, filename)
                
                messagebox.showinfo("Succès", f"Rapport exporté vers:\n{filename}")
                
            except Exception as e:
                messagebox.showerror("Erreur", f"Erreur lors de l'export: {e}")
    
    def export_to_pdf(self, report_type, filename):
        """Exporte vers PDF"""
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        
        doc = SimpleDocTemplate(filename, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Titre
        title = Paragraph(f"Rapport {report_type.title()}", styles['Title'])
        story.append(title)
        story.append(Spacer(1, 12))
        
        # Date de génération
        date_para = Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", styles['Normal'])
        story.append(date_para)
        story.append(Spacer(1, 12))
        
        if report_type == 'banks':
            # Données des banques
            data = [['Banque', 'Nb Chèques', 'Montant Total', 'Montant Moyen']]
            
            for item in self.banks_tree.get_children():
                values = self.banks_tree.item(item)['values']
                data.append([values[0], values[1], values[2], values[3]])
            
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
        
        # Construire le PDF
        doc.build(story)
    
    def export_to_excel(self, report_type, filename):
        """Exporte vers Excel"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f"Rapport {report_type.title()}"
        
        if report_type == 'banks':
            # En-têtes
            headers = ['Banque', 'Nb Chèques', 'Montant Total', 'Montant Moyen', 'Encaissés', 'Rejetés', 'Part %']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Données
            for row, item in enumerate(self.banks_tree.get_children(), 2):
                values = self.banks_tree.item(item)['values']
                for col, value in enumerate(values, 1):
                    sheet.cell(row=row, column=col, value=value)
        
        workbook.save(filename)
    
    def export_to_csv(self, report_type, filename):
        """Exporte vers CSV"""
        import csv
        
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile, delimiter=';')
            
            if report_type == 'banks':
                # En-têtes
                writer.writerow(['Banque', 'Nb Chèques', 'Montant Total', 'Montant Moyen', 'Encaissés', 'Rejetés', 'Part %'])
                
                # Données
                for item in self.banks_tree.get_children():
                    values = self.banks_tree.item(item)['values']
                    writer.writerow(values)
    
    def email_report(self):
        """Envoie le rapport par email"""
        messagebox.showinfo("Email", "Fonctionnalité d'envoi par email à implémenter")
    
    def generate_monthly_report(self):
        """Génère un rapport mensuel"""
        self.notebook.select(0)  # Onglet résumé
        
        # Définir la période du mois en cours
        today = date.today()
        first_day = today.replace(day=1)
        
        self.summary_date_from.set_date(first_day)
        self.summary_date_to.set_date(today)
        
        self.generate_summary()
    
    def show_bank_analysis(self):
        """Affiche l'analyse par banque"""
        self.notebook.select(2)  # Onglet banques
        self.update_banks_analysis()
